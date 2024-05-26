import numpy as np
import matplotlib.pyplot as plt

class FuzzyDEMATELSolver:
    def __init__(self):
        self.experts = []
        self.factors = []
        self.numberOfExperts = 0
        self.numberOfFactors = 0
        
        
    def setNumberOfExperts(self, nb):
        self.numberOfExperts=nb
    def getNumberOfExperts(self):
        return self.numberOfExperts   
    
    def setNumberOfFactors(self, nb):
        self.numberOfFactors=nb
    def getNumberOfFactors(self):
        return self.numberOfFactors  

    
    def getMatrix(self):
        return self.matrix
    def AddMatrix(self, matrix):
        self.matrix.append(matrix)       
    def setMatrix(self, matrices):
        self.matrix=matrices
    def setExperts(self, experts):
            self.experts=experts
    def getExperts(self):
        return self.experts            
    def addExpert(self, expert):
            self.experts.append(expert)
    def setFactors(self, factors):
                self.factors=factors  
    def getFactors(self):
        return self.factors                   
    def addFactor(self, factor):
                self.factors.append(factor)
    def getFuzzyDirectInfluenceMatrix(self):
        return self.Z
          
    def getFuzzyNormalizedDirectInfluenceMatrix (self):
        return self.X
        
    def getFuzzyTotalInfluenceMatrix (self):
        return self.T
     
    def getProminence(self):
        return self.R + self.C
    
    def getRalation(self):
        return self.R - self.C
    
    
    def setNumberOfExperts(self, nb):
        self.numberOfExperts = nb
    
    def setNumberOfFactors(self, nb):
        self.numberOfFactors = nb
 
    
    def step1(self):

        dims = self.matrix[0].shape
        self.Z  = np.zeros(dims)
        self.matrix = self.matrix.astype(np.float64)
        for m in self.matrix:
            self.Z += m
        self.Z /= self.getNumberOfExperts() 
        
    def calculateS(self):
        U = self.Z[:, :, 2]
        sumCol = np.sum(U, axis=0)
        sumRow = np.sum(U, axis=1)
        sup=max(max(sumRow),max(sumCol))
        return sup
    def step2(self):
        self.X = self.Z / self.calculateS()  
    def step3(self):
        I = np.eye(self.numberOfFactors)

        XL = self.X[:, :, 0]
        XM = self.X[:, :, 1]
        XU = self.X[:, :, 2]
        TL=np.dot(XL,np.linalg.inv(I-XL))
        TM=np.dot(XM,np.linalg.inv(I-XM))
        TU=np.dot(XU,np.linalg.inv(I-XU))
        self.T = np.empty((self.getNumberOfFactors(), self.getNumberOfFactors(), 3), dtype=TL.dtype)
        
        self.T[:, :, 0] = TL
        self.T[:, :, 1] = TM
        self.T[:, :, 2] = TU

    def step4(self):
        self.FR = np.empty((self.T.shape[0], 3), dtype=self.T.dtype)
        self.FC = np.empty((self.T.shape[0], 3), dtype=self.T.dtype)
        prominence=np.empty((self.T.shape[0], 3), dtype=self.T.dtype)
        relation=np.empty((self.T.shape[0], 3), dtype=self.T.dtype)

        # Calculate the sums on index j for each element of T and assign them to each element of R
        self.FR[:, 0] = np.sum(self.T[:, :, 0], axis=1)
        self.FR[:, 1] = np.sum(self.T[:, :, 1], axis=1)
        self.FR[:, 2] = np.sum(self.T[:, :, 2], axis=1)
        # Calculate the sums on index j for each element of T and assign them to each element of R 
        self.FC[:, 0] = np.sum(self.T[:, :, 0], axis=0)
        self.FC[:, 1] = np.sum(self.T[:, :, 1], axis=0)
        self.FC[:, 2] = np.sum(self.T[:, :, 2], axis=0)
        self.FProminence=self.FR+self.FC
        self.FRelation=self.FR-self.FC

        self.R=np.zeros(self.getNumberOfFactors(), dtype=self.T.dtype)
        self.C=np.zeros(self.getNumberOfFactors(), dtype=self.T.dtype)
        for i in range(self.getNumberOfFactors()):
            self.R[i]=((self.FR[i][2]-self.FR[i][0])+(self.FR[i][1]-self.FR[i][0]))/3 + self.FR[i][0]
            self.C[i]=((self.FC[i][2]-self.FC[i][0])+(self.FC[i][1]-self.FC[i][0]))/3 + self.FC[i][0]
        self.result={"cause":[],"effect":[]}
        self.prominence=self.getProminence()
        self.relation=self.getRalation()

        for i in range(len(self.relation)):
            if(self.relation[i]<0): self.result["effect"].append(self.factors[i])
            else : self.result["cause"].append(self.factors[i])

    def drawCurve(self):
            fig, ax = plt.subplots(figsize=(2, 2)) 
            ax.scatter(self.prominence,  self.relation, marker='o', s=50)  
            for name,p,r in zip(self.factors, self.prominence, self.relation):
                ax.text(p, r, name, ha='center', va='bottom')
            ax.set_xlabel('R+C')
            ax.set_ylabel('R-C')
            ax.axhline(0, color='black', linestyle='--')    
    def savexl(self,url):
        from openpyxl import Workbook
        wb = Workbook()
        ws5 = wb.active
        ws5.title = "Defuzzified IRM"
        ws5['A1'] = "Factor name"
        ws5['B1'] = "R"
        ws5['C1'] = "C"
        ws5['D1'] = "R+C"
        ws5['E1'] = "R-C"
        i=2
        for nameFactor,r,c,rpc,rmc in zip(self.factors,self.R,self.C,self.prominence,self.relation):
             ws5['A'+str(i)] = nameFactor
             ws5['B'+str(i)] = r
             ws5['C'+str(i)] = c
             ws5['D'+str(i)] = rpc
             ws5['E'+str(i)] = rmc
             i+=1
        ws1 = wb.create_sheet(title="Fuzzy Influential Relation Map")
        ws1.title = " IRM"
        ws1['A1'] = "Factor name"
        ws1['B1'] = "R"
        ws1['C1'] = "C"
        ws1['D1'] = "R+C"
        ws1['E1'] = "R-C"
        i=2
        for nameFactor,r,c,rpc,rmc in zip(self.factors,self.FR,self.FC,self.FProminence,self.FRelation):
             ws1['A'+str(i)] = nameFactor
             ws1['B'+str(i)] = "("+str(r[0])+","+str(r[1])+","+str(r[2])+")"
             ws1['C'+str(i)] = "("+str(c[0])+","+str(c[1])+","+str(c[2])+")"
             ws1['D'+str(i)] = "("+str(rpc[0])+","+str(rpc[1])+","+str(rpc[2])+")"
             ws1['E'+str(i)] = "("+str(rmc[0])+","+str(rmc[1])+","+str(rmc[2])+")"
             i+=1

        ws2 = wb.create_sheet(title="Direct-influence fuzzy matrix")
        i=2
        for nameFactor in self.factors:
             ws2['A'+str(i)] = nameFactor
             coord=chr(ord('A')+i-1)+'1'
             ws2[coord]=nameFactor
             i+=1
        for i in range(2, self.numberOfFactors + 2):
            for j in range(2, self.numberOfFactors + 2):
                lettre=chr(ord('B')+j-2)
                num=str(i)
                ws2[lettre+num] = "("+str(self.Z[i-2, j-2] [0])+","+str(self.Z[i-2, j-2] [1])+","+str(self.Z[i-2, j-2] [2])+")" 
        ws3 = wb.create_sheet(title="NormIinfluence Fuzzy matrix")
        i=2
        for nameFactor in self.factors:
             ws3['A'+str(i)] = nameFactor
             coord=chr(ord('A')+i-1)+'1'
             ws3[coord]=nameFactor
             i+=1
        for i in range(2, self.numberOfFactors + 2):
            for j in range(2, self.numberOfFactors + 2):
                lettre=chr(ord('B')+j-2)
                num=str(i)
                ws3[lettre+num] = "("+str(self.X[i-2, j-2][0])+","+str(self.X[i-2, j-2][1])+","+str(self.X[i-2, j-2][2])+")"     
        ws4 = wb.create_sheet(title="Total-influence Fuzzy matrix")
        i=2
        for nameFactor in self.factors:
             ws4['A'+str(i)] = nameFactor
             coord=chr(ord('A')+i-1)+'1'
             ws4[coord]=nameFactor
             i+=1
        for i in range(2, self.numberOfFactors + 2):
            for j in range(2, self.numberOfFactors + 2):
                lettre=chr(ord('B')+j-2)
                num=str(i)
                ws4[lettre+num] = "("+str(self.T[i-2, j-2][0])+","+str(self.T[i-2, j-2][1])+","+str(self.T[i-2, j-2][2])+")"    
        wb.save(url+"/DEMATELAnalysis.xlsx")
        print("Fichier Excel créé avec succès.")

    def     testFuction(self):
            # Exemple d'utilisation
            expert = "Alami"
            factors = [ "A1 ", 
                        "A2 ", 
                        "A3 ", 
                        "A4 ", 
                        "A5 ", 
                    "A6 ", 
                        "A7 ", 
                    "A8",
                    "A9",
                    "A10",           
                    "A11", 
                        "A12 ", 
                        "A13 ", 
                        "A14 ", 
                        "A15 ", 
                    "A16 ", 
                        "A17 ", 
                    "A18"]           
            # example source https://www.sciencedirect.com/science/article/pii/S0950423015300498
            valeurs=[
            [(0, 0, 0.25),(0, 0, 0.25),(0.5, 0.75, 1),(0.25, 0.5, 0.75),(0, 0.25, 0.5),(0, 0.25, 0.5),(0, 0, 0.25),(0, 0, 0.25),(0, 0.25, 0.5),(0.5, 0.75, 1),(0.25, 0.5, 0.75),(0, 0.25, 0.5),(0, 0, 0.25),(0, 0, 0.25),(0.25, 0.5, 0.75),(0, 0, 0.25),(0, 0.25, 0.5),(0.5, 0.75, 1)],
            [(0.5, 0.75, 1),(0, 0, 0.25),(0.5, 0.75, 1),(0.25, 0.5, 0.75),(0, 0, 0.25),(0.25, 0.5, 0.75),(0, 0.25, 0.5),(0, 0.25, 0.5),(0, 0, 0.25),(0, 0.25, 0.5),(0.75, 1, 1),(0, 0.25, 0.5),(0, 0, 0.25),(0, 0.25, 0.5),(0.25, 0.5, 0.75),(0, 0.25, 0.5),(0, 0.25, 0.5),(0.5, 0.75, 1)],
            [(0, 0.25, 0.5),(0.25, 0.5, 0.75),(0, 0, 0.25),(0.5, 0.75, 1),(0.25, 0.5, 0.75),(0, 0.25, 0.5),(0.25, 0.5, 0.75),(0, 0, 0.25),(0, 0.25, 0.5),(0.25, 0.5, 0.75),(0, 0, 0.25),(0, 0.25, 0.5),(0.25, 0.5, 0.75),(0, 0.25, 0.5),(0.5, 0.75, 1),(0, 0, 0.25),(0.25, 0.5, 0.75),(0.25, 0.5, 0.75)],
            [(0.5, 0.75, 1),(0.25, 0.5, 0.75),(0, 0.25, 0.5),(0, 0, 0.25),(0.25, 0.5, 0.75),(0.5, 0.75, 1),(0.75, 1, 1),(0, 0.25, 0.5),(0.25, 0.5, 0.75),(0.5, 0.75, 1),(0.25, 0.5, 0.75),(0.75, 1, 1),(0.5, 0.75, 1),(0, 0.25, 0.5),(0.5, 0.75, 1),(0, 0.25, 0.5),(0, 0, 0.25),(0, 0.25, 0.5)],
            [(0, 0.25, 0.5),(0, 0.25, 0.5),(0.5, 0.75, 1),(0.25, 0.5, 0.75),(0, 0, 0.25),(0.5, 0.75, 1),(0.5, 0.75, 1),(0, 0, 0.25),(0, 0.25, 0.5),(0, 0.25, 0.5),(0.75, 1, 1),(0.5, 0.75, 1),(0, 0.25, 0.5),(0.25, 0.5, 0.75),(0, 0.25, 0.5),(0.25, 0.5, 0.75),(0, 0, 0.25),(0.5, 0.75, 1)],
            [(0, 0.25, 0.5),(0.25, 0.5, 0.75),(0.25, 0.5, 0.75),(0, 0.25, 0.5),(0, 0.25, 0.5),(0, 0, 0.25),(0.25, 0.5, 0.75),(0, 0, 0.25),(0, 0.25, 0.5),(0, 0, 0.25),(0, 0.25, 0.5),(0, 0.25, 0.5),(0.25, 0.5, 0.75),(0.25, 0.5, 0.75),(0, 0, 0.25),(0, 0.25, 0.5),(0, 0.25, 0.5),(0.75, 1, 1)],
            [(0, 0, 0.25),(0, 0.25, 0.5),(0.25, 0.5, 0.75),(0.5, 0.75, 1),(0.25, 0.5, 0.75),(0.25, 0.5, 0.75),(0, 0, 0.25),(0, 0.25, 0.5),(0.25, 0.5, 0.75),(0, 0.25, 0.5),(0.25, 0.5, 0.75),(0, 0.25, 0.5),(0.25, 0.5, 0.75),(0, 0.25, 0.5),(0, 0, 0.25),(0, 0.25, 0.5),(0.25, 0.5, 0.75),(0.25, 0.5, 0.75)],
            [(0, 0, 0.25),(0, 0.25, 0.5),(0, 0.25, 0.5),(0, 0.25, 0.5),(0.25, 0.5, 0.75),(0, 0, 0.25),(0, 0, 0.25),(0, 0, 0.25),(0.5, 0.75, 1),(0.5, 0.75, 1),(0, 0, 0.25),(0, 0.25, 0.5),(0, 0.25, 0.5),(0, 0, 0.25),(0, 0.25, 0.5),(0, 0.25, 0.5),(0, 0, 0.25),(0, 0, 0.25)],
            [(0, 0, 0.25),(0, 0.25, 0.5),(0.25, 0.5, 0.75),(0, 0.25, 0.5),(0.25, 0.5, 0.75),(0, 0, 0.25),(0, 0, 0.25),(0.5, 0.75, 1),(0, 0, 0.25),(0.25, 0.5, 0.75),(0, 0.25, 0.5),(0, 0, 0.25),(0, 0, 0.25),(0, 0.25, 0.5),(0, 0.25, 0.5),(0, 0, 0.25),(0.25, 0.5, 0.75),(0, 0, 0.25)],
            [(0, 0.25, 0.5),(0.25, 0.5, 0.75),(0, 0, 0.25),(0.5, 0.75, 1),(0, 0.25, 0.5),(0, 0.25, 0.5),(0, 0.25, 0.5),(0.25, 0.5, 0.75),(0.75, 1, 1),(0, 0, 0.25),(0, 0.25, 0.5),(0.5, 0.75, 1),(0, 0.25, 0.5),(0, 0, 0.25),(0, 0.25, 0.5),(0.25, 0.5, 0.75),(0, 0.25, 0.5),(0.5, 0.75, 1)],
            [(0, 0.25, 0.5),(0.25, 0.5, 0.75),(0.5, 0.75, 1),(0.75, 1, 1),(0, 0.25, 0.5),(0.25, 0.5, 0.75),(0.5, 0.75, 1),(0, 0, 0.25),(0, 0.25, 0.5),(0.5, 0.75, 1),(0, 0, 0.25),(0.75, 1, 1),(0, 0.25, 0.5),(0, 0, 0.25),(0, 0.25, 0.5),(0, 0, 0.25),(0, 0.25, 0.5),(0.5, 0.75, 1)],
            [(0.25, 0.5, 0.75),(0.5, 0.75, 1),(0.75, 1, 1),(0.5, 0.75, 1),(0, 0, 0.25),(0, 0.25, 0.5),(0.75, 1, 1),(0, 0, 0.25),(0, 0.25, 0.5),(0.5, 0.75, 1),(0, 0.25, 0.5),(0, 0, 0.25),(0, 0, 0.25),(0, 0.25, 0.5),(0.5, 0.75, 1),(0, 0, 0.25),(0, 0, 0.25),(0.5, 0.75, 1)],
            [(0.5, 0.75, 1),(0.25, 0.5, 0.75),(0.75, 1, 1),(0.25, 0.5, 0.75),(0.75, 1, 1),(0, 0.25, 0.5),(0.5, 0.75, 1),(0, 0, 0.25),(0, 0, 0.25),(0, 0.25, 0.5),(0.5, 0.75, 1),(0.5, 0.75, 1),(0, 0, 0.25),(0.5, 0.75, 1),(0, 0.25, 0.5),(0, 0, 0.25),(0, 0.25, 0.5),(0, 0.25, 0.5)],
            [(0, 0.25, 0.5),(0, 0.25, 0.5),(0.5, 0.75, 1),(0.25, 0.5, 0.75),(0, 0.25, 0.5),(0, 0.25, 0.5),(0, 0, 0.25),(0, 0, 0.25),(0, 0, 0.25),(0, 0, 0.25),(0.25, 0.5, 0.75),(0, 0.25, 0.5),(0.25, 0.5, 0.75),(0, 0, 0.25),(0, 0, 0.25),(0, 0, 0.25),(0, 0, 0.25),(0.5, 0.75, 1)],
            [(0.25, 0.5, 0.75),(0.25, 0.5, 0.75),(0.5, 0.75, 1),(0.75, 1, 1),(0, 0.25, 0.5),(0.25, 0.5, 0.75),(0.5, 0.75, 1),(0, 0.25, 0.5),(0, 0, 0.25),(0, 0.25, 0.5),(0.25, 0.5, 0.75),(0, 0.25, 0.5),(0.5, 0.75, 1),(0.5, 0.75, 1),(0, 0, 0.25),(0.75, 1, 1),(0.25, 0.5, 0.75),(0.5, 0.75, 1)],
            [(0, 0.25, 0.5),(0.75, 1, 1),(0.5, 0.75, 1),(0.25, 0.5, 0.75),(0, 0.25, 0.5),(0, 0.25, 0.5),(0, 0.25, 0.5),(0, 0, 0.25),(0, 0, 0.25),(0, 0.25, 0.5),(0.25, 0.5, 0.75),(0.25, 0.5, 0.75),(0, 0, 0.25),(0, 0.25, 0.5),(0.25, 0.5, 0.75),(0, 0, 0.25),(0.5, 0.75, 1),(0.75, 1, 1)],
            [(0, 0, 0.25),(0.5, 0.75, 1),(0.75, 1, 1),(0.5, 0.75, 1),(0, 0.25, 0.5),(0, 0, 0.25),(0, 0.25, 0.5),(0, 0, 0.25),(0, 0.25, 0.5),(0, 0.25, 0.5),(0, 0.25, 0.5),(0, 0.25, 0.5),(0, 0, 0.25),(0, 0.25, 0.5),(0, 0.25, 0.5),(0.25, 0.5, 0.75),(0, 0, 0.25),(0.25, 0.5, 0.75)],
            [(0, 0.25, 0.5),(0.5, 0.75, 1),(0.75, 1, 1),(0.25, 0.5, 0.75),(0, 0.25, 0.5),(0, 0.25, 0.5),(0, 0, 0.25),(0, 0, 0.25),(0, 0.25, 0.5),(0.25, 0.5, 0.75),(0.5, 0.75, 1),(0.25, 0.5, 0.75),(0, 0.25, 0.5),(0, 0, 0.25),(0.25, 0.5, 0.75),(0.5, 0.75, 1),(0, 0.25, 0.5),(0, 0, 0.25)]
            ]

            matrices = np.array([np.array(valeurs, dtype=object)])
            # Creating a DEMATELSolver instance
            solver = FuzzyDEMATELSolver()
            solver.setMatrix(matrices)
            print(solver.getMatrix())
            solver.addExpert(expert)
            solver.setFactors(factors)

            self.setMatrix(matrices)
            self.addExpert(expert)
            self.setFactors(factors)

            self.setNumberOfExperts(1)
            self.setNumberOfFactors(len(solver.factors))

            print("Factor list :\n",self.getFactors())

            self.step1()
            print("Direct Influence Fuzzy Matrix :\n",self.getFuzzyDirectInfluenceMatrix())
            self.step2()
            print("Normalized Direct Influence Fuzzy Matrix :\n",+self.getFuzzyNormalizedDirectInfluenceMatrix())
            self.step3()
            print("Total Influence Fuzzy Matrix :\n",self.getFuzzyTotalInfluenceMatrix())
            self.step4()
            print("Relation :\n",self.getRalation())
            print("Prominence :\n",self.getProminence())
            self.drawCurve()
            self.savexl(input("Please enter the destination path for the Excel file:"))
             



