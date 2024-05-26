import numpy as np
import matplotlib.pyplot as plt

class DEMATELSolver:
    def __init__(self):
        self.experts = []
        self.factors = []
        self.numberOfExperts = 0
        self.numberOfFactors = 0
        
    def setNumberOfExperts(self, nb):
        self.numberOfExperts=nb
    def getNumberOfExperts(self):
        return self.numberOfExperts   
    
    def setNumberOfFactors(self, nb):
        self.numberOfFactors=nb
    def getNumberOfFactors(self):
        return self.numberOfFactors  

    
    def getMatrix(self):
        return self.matrix
    def setMatrix(self, matrices):
        self.matrix=matrices
    def AddMatrix(self, matrix):
        self.matrix.append(matrix)   
    def setExperts(self, experts):
            self.experts=experts
    def getExperts(self):
        return self.experts            
    def addExpert(self, expert):
            self.experts.append(expert)
    def setFactors(self, factors):
                self.factors=factors  
    def getFactors(self):
        return self.factors                   
    def addFactor(self, factor):
                self.factors.append(factor)
    def getDirectInfluenceMatrix(self):
        return self.Z
          
    def getNormalizedDirectInfluenceMatrix (self):
        return self.X
        
    def getTotalInfluenceMatrix (self):
        return self.T
     
    def getProminence(self):
        return self.R + self.C
    
    def getRalation(self):
        return self.R - self.C
    
    
    def setNumberOfExperts(self, nb):
        self.numberOfExperts = nb
    
    def setNumberOfFactors(self, nb):
        self.numberOfFactors = nb
    
    def step1(self):
        self.Z = np.zeros((self.getNumberOfFactors(), self.getNumberOfFactors()))
        for matrix in self.matrix:
            self.Z += matrix
        self.Z /= self.getNumberOfExperts()
        self.Z=self.Z  
        
    
    def step2(self):
        self.X = self.Z / self.calculateS()

         
    def step3(self):
        I = np.eye(self.getNumberOfFactors())
        self.T=np.dot(self.X,np.linalg.inv(I-self.X))
        

    def step4(self):
        self.result={"cause":[],"effect":[]}
        self.R = np.sum(self.T, axis=1)
        self.C = np.sum(self.T, axis=0)
        self.prominence=self.getProminence()
        self.relation=self.getRalation()
        for i in range(len(self.relation)):
            if(self.relation[i]<0): self.result["effect"].append(self.factors[i])
            else : self.result["cause"].append(self.factors[i])

    def calculateS(self):
        R = np.sum(self.Z, axis=1)
        C = np.sum(self.Z, axis=0)
        return max(np.max(R), np.max(C))

    def drawCurve(self):
            fig, ax = plt.subplots(figsize=(2, 2)) 
            ax.scatter(self.prominence,  self.relation, marker='o', s=50)  
            for name,p,r in zip(self.factors, self.prominence, self.relation):
                ax.text(p, r, name, ha='center', va='bottom')
            ax.set_xlabel('R+C')
            ax.set_ylabel('R-C')
            ax.axhline(0, color='black', linestyle='--')
    def savexl(self,url):
        from openpyxl import Workbook
        wb = Workbook()
        ws1 = wb.active
        ws1.title = "Influential Relation Map IRM"
        ws1['A1'] = "Factor name"
        ws1['B1'] = "R"
        ws1['C1'] = "C"
        ws1['D1'] = "R+C"
        ws1['E1'] = "R-C"
        i=2
        for nameFactor,r,c,rpc,rmc in zip(self.factors,self.R,self.C,self.prominence,self.relation):
             ws1['A'+str(i)] = nameFactor
             ws1['B'+str(i)] = r
             ws1['C'+str(i)] = c
             ws1['D'+str(i)] = rpc
             ws1['E'+str(i)] = rmc
             i+=1
        ws2 = wb.create_sheet(title="Direct-influence matrix")
        i=2
        for nameFactor in self.factors:
             ws2['A'+str(i)] = nameFactor
             coord=chr(ord('A')+i-1)+'1'
             ws2[coord]=nameFactor
             i+=1
        for i in range(2, self.numberOfFactors + 2):
            for j in range(2, self.numberOfFactors + 2):
                lettre=chr(ord('B')+j-2)
                num=str(i)
                ws2[lettre+num] = self.Z[i-2, j-2] 
        ws3 = wb.create_sheet(title="Normalized Iinfluence matrix")
        i=2
        for nameFactor in self.factors:
             ws3['A'+str(i)] = nameFactor
             coord=chr(ord('A')+i-1)+'1'
             ws3[coord]=nameFactor
             i+=1
        for i in range(2, self.numberOfFactors + 2):
            for j in range(2, self.numberOfFactors + 2):
                lettre=chr(ord('B')+j-2)
                num=str(i)
                ws3[lettre+num] = self.X[i-2, j-2] 
        ws4 = wb.create_sheet(title="Total-influence matrix")
        i=2
        for nameFactor in self.factors:
             ws4['A'+str(i)] = nameFactor
             coord=chr(ord('A')+i-1)+'1'
             ws4[coord]=nameFactor
             i+=1
        for i in range(2, self.numberOfFactors + 2):
            for j in range(2, self.numberOfFactors + 2):
                lettre=chr(ord('B')+j-2)
                num=str(i)
                ws4[lettre+num] = self.T[i-2, j-2] 
        wb.save(url+"/DEMATELAnalysis.xlsx")
    def testFuction(self):
            # Exemple d'utilisation
            expert = "Alami"
            factors = [ "A1 ", 
                        "A2 ", 
                        "A3 ", 
                        "A4 ", 
                        "A5 ", 
                    "A6 ", 
                        "A7 ", 
                    "A8"]

            # Direct influence matrices for each expert (example)
            matrices = [np.array([[0.00, 1.11, 1.01, 1.41, 1.66, 0.50, 1.60, 2.00],                   
                                [1.43, 0.00, 2.22, 2.00, 2.40, 1.20, 1.66, 1.33], 
                                [0.82, 1.00, 0.00, 2.05, 2.44, 1.65, 2.65, 1.88], 
                                [1.92, 0.80, 1.82, 0.00, 2.75, 3.50, 3.33, 3.10], 
                                [2.20, 3.11, 1.25, 0.75, 0.00, 2.25, 2.66, 1.75], 
                                [1.01, 1.31, 1.45, 1.20, 1.44, 0.00, 0.75, 3.00], 
                                [3.50, 3.20, 2.95, 3.33, 2.88, 1.85, 0.00, 1.40], 
                                [0.50, 1.25, 1.40, 3.66, 1.00, 2.00, 3.33, 0.00]])]
            # Creating a DEMATELSolver instance
            self = DEMATELSolver()
            self.setMatrix(matrices)
            self.addExpert(expert)
            self.setFactors(factors)

            self.setNumberOfExperts(1)
            self.setNumberOfFactors(8)

            print("Factor list :\n",self.getFactors())
            # Exécution des étapes
            self.step1()
            print("Direct Influence Matrix :\n",self.getDirectInfluenceMatrix())
            self.step2()
            print("Normalized Direct Influence Matrix :\n",+self.getNormalizedDirectInfluenceMatrix())
            self.step3()
            print("Total Influence Matrix :\n",self.getTotalInfluenceMatrix())
            self.step4()
            print("Relation :\n",self.getRalation())
            print("Prominence :\n",self.getProminence())
            self.drawCurve()
            self.savexl(input("Please enter the destination path for the Excel file:"))
             




